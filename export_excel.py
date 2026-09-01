import sqlite3
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_duration_in_days(start_str, end_str, year=2026):
    """
    Calculates the number of days a vacation spans inside a given year.
    """
    try:
        start_date = datetime.datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_str, "%Y-%m-%d").date()
        
        year_start = datetime.date(year, 1, 1)
        year_end = datetime.date(year, 12, 31)
        
        actual_start = max(start_date, year_start)
        actual_end = min(end_date, year_end)
        
        if actual_start <= actual_end:
            return (actual_end - actual_start).days + 1
        return 0
    except Exception:
        return 0

def export_to_excel(db_path="vacations.db", output_path="vacations_report.xlsx", year=2026):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Fetch employees
    cursor.execute("SELECT id, name, initial, color, main_vacation_limit, additional_vacation_limit, other_vacation_limit FROM employees")
    employees = cursor.fetchall()
    
    # Fetch vacations
    cursor.execute("SELECT employee_id, start_date, end_date, type, status FROM vacations")
    vacations = cursor.fetchall()
    
    conn.close()
    
    # Group vacations by employee ID
    vacations_by_emp = {}
    for vac in vacations:
        emp_id, start_date, end_date, v_type, status = vac
        if emp_id not in vacations_by_emp:
            vacations_by_emp[emp_id] = []
        vacations_by_emp[emp_id].append({
            "start": start_date,
            "end": end_date,
            "type": v_type,
            "status": status
        })
        
    wb = Workbook()
    ws = wb.active
    ws.title = f"Відпустки {year}"
    
    # Styling
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, bold=False)
    bold_body_font = Font(name="Arial", size=10, bold=True)
    
    title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    accent_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"ЗВІТ ПО ВІДПУСТКАХ ПРАЦІВНИКІВ ЗА {year} РІК"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = [
        "ПІБ Працівника", "Ініціали", 
        "Основна (Ліміт)", "Основна (Вик.)", "Основна (Залишок)",
        "Додаткова (Ліміт)", "Додаткова (Вик.)", "Додаткова (Залишок)",
        "Інші (Ліміт)", "Інші (Вик.)", "Інші (Залишок)",
        "Періоди відпусток"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[2].height = 28
    
    # Write data
    current_row = 3
    for emp in employees:
        emp_id, name, initial, color, main_lim, add_lim, oth_lim = emp
        emp_vacs = vacations_by_emp.get(emp_id, [])
        
        # Calculate used days for each type in the given year
        main_used = sum(get_duration_in_days(v["start"], v["end"], year) for v in emp_vacs if v["type"] == "основна")
        add_used = sum(get_duration_in_days(v["start"], v["end"], year) for v in emp_vacs if v["type"] == "додаткова")
        oth_used = sum(get_duration_in_days(v["start"], v["end"], year) for v in emp_vacs if v["type"] == "інша")
        
        main_rem = main_lim - main_used
        add_rem = add_lim - add_used
        oth_rem = oth_lim - oth_used
        
        # Vacation periods text
        periods_text = "; ".join(
            f"{v['start']} - {v['end']} ({v['type']}, {v['status']})" for v in emp_vacs
        )
        
        row_values = [
            name, initial,
            main_lim, main_used, main_rem,
            add_lim, add_used, add_rem,
            oth_lim, oth_used, oth_rem,
            periods_text
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = body_font
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 12]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
            # Formatting/Highlighting
            if col_idx in [5, 8, 11] and isinstance(val, (int, float)) and val < 0:
                # Overlimit in red
                cell.font = Font(name="Arial", size=10, bold=True, color="FF0000")
                
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                # Limit length calculation to avoid overly wide column for periods
                val_str = str(cell.value)
                if len(val_str) > 30:
                    max_len = max(max_len, 30)
                else:
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
    print(f"Звіт збережено успішно: {output_path}")

if __name__ == "__main__":
    export_to_excel()
